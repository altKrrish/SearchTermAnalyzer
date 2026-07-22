"""
Amazon Search Term & ASIN Analyzer — Production Flask Web App for Render
Stateless, Atomic File Processing & High-Performance Excel Analytics Engine
"""

import os
import re
import time
import uuid
import tempfile
from collections import defaultdict

from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl.utils

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB limit
app.config['PROPAGATE_EXCEPTIONS'] = False

OUTPUT_DIR = os.path.join(tempfile.gettempdir(), "sta_outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ---------------------------------------------------------------------------
# Global Exception Handler (Guarantees JSON error response)
# ---------------------------------------------------------------------------
@app.errorhandler(Exception)
def handle_unexpected_error(e):
    import traceback
    code = 500
    if hasattr(e, 'code') and isinstance(e.code, int):
        code = e.code
    return jsonify({
        "error": str(e),
        "trace": traceback.format_exc()
    }), code


# ---------------------------------------------------------------------------
# Classification rules  (order matters – first match wins)
# ---------------------------------------------------------------------------
CLASSIFICATION_RULES = [
    ("Protective White Ruffle Pillow Covers", "Pillow Covers", [r"protective.*white.*ruffle.*pillow", r"white.*ruffle.*protective.*pillow"]),
    ("Ruffle Bedding Pillow Covers", "Pillow Covers", [r"ruffle.*bedding.*pillow", r"bedding.*ruffle.*pillow"]),
    ("Hand Print Pillow Covers", "Pillow Covers", [r"hand\s*print.*pillow", r"hand.*printed.*pillow"]),
    ("Piping Sham Cover", "Pillow Covers", [r"piping.*sham"]),
    ("Ruffle Sham Cover", "Pillow Covers", [r"ruffle.*sham"]),
    ("Piping Pillow Cover", "Pillow Covers", [r"piping.*pillow"]),
    ("Ruffle Pillow Cover", "Pillow Covers", [r"ruffle.*pillow", r"pillow.*ruffle"]),
    ("Piping Sham Cover", "Pillow Covers", [r"\bsham\b"]),
    ("Pillow Covers", "Pillow Covers", [r"pillow", r"cushion", r"throw", r"lumbar", r"outdoor.*pillow"]),
    ("Round Tablecloth", "Tablecloths", [r"\bround\b"]),
    ("Square Tablecloth", "Tablecloths", [r"\bsquare\b", r"\bcard\s*table\b", r"\bmahjong\b", r"54\s*x\s*54", r"60\s*x\s*60", r"72\s*x\s*72"]),
    ("Rectangle Tablecloth", "Tablecloths", [r"\brectangle\b", r"\brectangular\b", r"60\s*x\s*120", r"60\s*x\s*84", r"60\s*x\s*90", r"144\s*x\s*60", r"\b6\s*foot\b", r"\b8\s*foot\b"]),
    ("Rectangle Tablecloth", "Tablecloths", [r"tablecloth", r"table\s*cloth"]),
    ("Spa Towel", "Towels", [r"\bspa\b", r"\bbath\b"]),
    ("Dish Cloth", "Towels", [r"\bdish\b"]),
    ("Kitchen Towel", "Towels", [r"kitchen", r"tea\s*towel", r"hand\s*towel", r"waffle", r"towel"]),
    ("Curtains", "Curtains", [r"."]),
    ("Napkins", "Napkins", [r"."]),
    ("Placemats", "Placemats", [r"."]),
    ("Table Runner", "Table Runners", [r"."]),
    ("Tissue Box Covers", "Other", [r"tissue", r"kleenex"]),
    ("Lampshades", "Other", [r"lamp\s*shade", r"lampshade"]),
    ("Seat Cushion", "Other", [r"seat\s*cushion", r"chair\s*cushion", r"cushion\s*cover"]),
    ("Aprons", "Other", [r"\bapron"]),
    ("Spa Towel", "Other", [r"\bspa\b.*towel", r"towel.*\bspa\b"]),
    ("Kitchen Towel", "Other", [r"kitchen\s*towel", r"tea\s*towel"]),
    ("Dish Cloth", "Other", [r"dish\s*cloth"]),
    ("Curtains", "Other", [r"curtain"]),
    ("Quilts & Bedspreads", "Quilts & Bedspreads", [r"."]),
    ("Coasters", "Coasters", [r"."]),
]

COLORS = [
    "light steel blue", "steel blue", "columbia blue", "cornflower blue", "cornflower",
    "navy blue", "navy", "sky blue", "blue",
    "asparagus green", "russian green", "sage green", "sage", "olive green", "olive", "green",
    "bubblegum pink", "pink", "rose",
    "mustard yellow", "mustard", "yellow",
    "red", "maroon", "burgundy",
    "orange", "peach",
    "purple", "lavender", "lilac",
    "black", "white", "grey", "gray", "cream", "beige", "brown", "tan", "taupe", "teal"
]

ASIN_REGEX = re.compile(r'\bB0[A-Z0-9]{8}\b', re.IGNORECASE)


# ---------------------------------------------------------------------------
# Core logic functions
# ---------------------------------------------------------------------------
def parse_numeric(val):
    """Safely parse currency strings ($1,234.50), numbers, or percentages into float."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def extract_color(search_term):
    if not search_term:
        return None
    st_lower = str(search_term).lower()
    for color in COLORS:
        if re.search(r'\b' + re.escape(color) + r'\b', st_lower):
            return color.title()
    return None


def classify_search_term(broad_category, search_term):
    if not search_term:
        return "Uncategorized"
    st_lower = str(search_term).lower().strip()
    for subcat, cat_match, patterns in CLASSIFICATION_RULES:
        if cat_match and cat_match != broad_category:
            continue
        for pat in patterns:
            if re.search(pat, st_lower):
                return subcat
    return broad_category if broad_category else "Uncategorized"


def extract_asins_from_text(text):
    if not text:
        return []
    return [m.upper() for m in ASIN_REGEX.findall(str(text))]


def detect_header_row(ws):
    """Scan first 15 rows and pick the row with the most non-empty columns."""
    best_row_idx = 1
    best_row_data = []
    max_cols = 0
    max_c = max(ws.max_column or 100, 100)
    
    # Read into a list first because read_only mode does not support multiple iterations
    first_rows = list(ws.iter_rows(min_row=1, max_row=15, max_col=max_c, values_only=True))
    
    for i, row in enumerate(first_rows, 1):
        raw = list(row)
        while raw and (raw[-1] is None or str(raw[-1]).strip() == ""):
            raw.pop()
        col_count = sum(1 for v in raw if v is not None and str(v).strip() != "")
        if col_count > max_cols and col_count >= 2:
            max_cols = col_count
            best_row_idx = i
            best_row_data = raw
            
    if max_cols == 0:
        for i, row in enumerate(first_rows, 1):
            raw = list(row)
            while raw and (raw[-1] is None or str(raw[-1]).strip() == ""):
                raw.pop()
            col_count = sum(1 for v in raw if v is not None and str(v).strip() != "")
            if col_count > max_cols:
                max_cols = col_count
                best_row_idx = i
                best_row_data = raw
                
    return best_row_idx, best_row_data


def build_headers_list(raw_headers):
    headers = []
    for idx, h in enumerate(raw_headers, 1):
        val = str(h).strip() if h is not None else ""
        headers.append(val if val else f"Unnamed: Column {idx}")
    return headers


def auto_pick_keyword_col(headers):
    for h in headers:
        if any(hint in h.lower() for hint in ["search term", "customer search term", "keyword", "query", "search_term"]):
            return h
    return headers[0] if headers else None


def auto_pick_portfolio_col(headers):
    for h in headers:
        if "portfolio" in h.lower():
            return h
    return "(None)"


def find_sales_column_index(headers):
    for idx, h in enumerate(headers):
        h_lower = str(h).lower()
        if any(hint in h_lower for hint in ["sales ($)", "total_sales", "total sales", "sales", "revenue"]):
            return idx
    return 3 if len(headers) > 3 else 0


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------
title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_font = Font(name="Calibri", size=11, bold=True, color="2F5496")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
worst_section_font = Font(name="Calibri", size=11, bold=True, color="8B0000")
worst_section_fill = PatternFill(start_color="FDDEDE", end_color="FDDEDE", fill_type="solid")
data_font = Font(name="Calibri", size=10)
money_fmt = '#,##0.00'
pct_fmt = '0.00%'
thin_border = Border(bottom=Side(style="thin", color="B4C6E7"))
col_widths = {1: 18, 2: 45, 3: 14, 4: 14, 5: 10, 6: 10, 7: 10, 8: 10, 9: 14, 10: 22, 11: 45}


def write_analysis_sheet(wb, sheet_name, title_text, title_color,
                         headers, cat_label, best_rows, worst_rows):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_new = wb.create_sheet(sheet_name)
    new_headers = headers + [cat_label]
    num_cols = len(new_headers)

    ws_new.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws_new.cell(row=1, column=1, value=title_text)
    title_cell.font = title_font
    title_cell.fill = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_new.row_dimensions[1].height = 30

    for col_idx, hdr in enumerate(new_headers, 1):
        cell = ws_new.cell(row=2, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_new.row_dimensions[2].height = 25
    ws_new.freeze_panes = "A3"

    current_row = 3

    def write_block(block_label, rows, s_font, s_fill):
        nonlocal current_row
        if not rows:
            return
        ws_new.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        divider = ws_new.cell(row=current_row, column=1, value=f"  ▎ {block_label}")
        try:
            font_color = s_font.color.rgb if s_font.color and s_font.color.rgb else "2F5496"
        except Exception:
            font_color = "2F5496"
        divider.font = Font(name="Calibri", size=12, bold=True, color=font_color)
        divider.fill = s_fill
        divider.alignment = Alignment(vertical="center")
        ws_new.row_dimensions[current_row].height = 28
        current_row += 1

        prev_cat = None
        for subcat, vals in rows:
            if subcat != prev_cat:
                if prev_cat is not None:
                    current_row += 1
                ws_new.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
                count = sum(1 for s, _ in rows if s == subcat)
                section_cell = ws_new.cell(row=current_row, column=1, value=f"  {subcat.upper()}  |  {count} terms shown")
                section_cell.font = s_font
                section_cell.fill = s_fill
                section_cell.alignment = Alignment(vertical="center")
                ws_new.row_dimensions[current_row].height = 22
                current_row += 1
                prev_cat = subcat

            for col_idx, val in enumerate(vals, 1):
                cell = ws_new.cell(row=current_row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                col_name = headers[col_idx - 1] if col_idx - 1 < len(headers) else ""
                if col_name in ("Spend ($)", "Sales ($)", "Total_Spend", "Total_Sales"):
                    cell.number_format = money_fmt
                elif col_name in ("ACOS", "Effective_ACOS", "Effective_CVR", "Avg_CTR"):
                    cell.number_format = pct_fmt
                elif col_name in ("ROAS", "Effective_ROAS"):
                    cell.number_format = '#,##0.00'

            sub_cell = ws_new.cell(row=current_row, column=num_cols, value=subcat)
            sub_cell.font = Font(name="Calibri", size=10, bold=True, color="2F5496")
            sub_cell.border = thin_border
            current_row += 1
        current_row += 1

    write_block("🏆 TOP 10 BEST PERFORMING", best_rows, section_font, section_fill)
    write_block("⚠️ TOP 10 WORST PERFORMING", worst_rows, worst_section_font, worst_section_fill)

    for col_idx, width in col_widths.items():
        if col_idx <= num_cols:
            ws_new.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws_new.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(num_cols)}{current_row - 1}"


def write_asin_sheet(wb, all_asins, asin_rows, portfolio_idx):
    sn = "Extracted_ASINs"
    if sn in wb.sheetnames:
        del wb[sn]
    wa = wb.create_sheet(sn)
    has_portfolio = portfolio_idx is not None
    asin_headers = ["#", "ASIN", "Source Search Term"]
    if has_portfolio:
        asin_headers.append("Portfolio Name")

    wa.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(asin_headers))
    t = wa.cell(row=1, column=1, value=f"Extracted ASINs  |  {len(all_asins)} unique")
    t.font = title_font
    t.fill = PatternFill(start_color="1B7340", end_color="1B7340", fill_type="solid")
    t.alignment = Alignment(horizontal="left", vertical="center")
    wa.row_dimensions[1].height = 30

    ahf = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    for ci, h in enumerate(asin_headers, 1):
        cell = wa.cell(row=2, column=ci, value=h)
        cell.font = header_font
        cell.fill = ahf
        cell.alignment = Alignment(horizontal="center", vertical="center")
    wa.row_dimensions[2].height = 25
    wa.freeze_panes = "A3"

    for idx, (src, asin, port) in enumerate(asin_rows, 1):
        rn = idx + 2
        wa.cell(row=rn, column=1, value=idx).font = data_font
        wa.cell(row=rn, column=1).border = thin_border
        ac = wa.cell(row=rn, column=2, value=asin)
        ac.font = Font(name="Calibri", size=10, bold=True, color="1B7340")
        ac.border = thin_border
        sc = wa.cell(row=rn, column=3, value=src)
        sc.font = data_font
        sc.border = thin_border
        if has_portfolio:
            pc = wa.cell(row=rn, column=4, value=port)
            pc.font = data_font
            pc.border = thin_border

    wa.column_dimensions['A'].width = 8
    wa.column_dimensions['B'].width = 16
    wa.column_dimensions['C'].width = 60
    if has_portfolio:
        wa.column_dimensions['D'].width = 30
    wa.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(asin_headers))}{len(asin_rows) + 2}"


def build_best_worst(groups, sales_idx, mode):
    best_rows = []
    worst_rows = []
    summary = []
    for cat in sorted(groups.keys()):
        items = groups[cat]
        if mode == "ASIN":
            asin_best = {}
            for asin, vals in items:
                raw_val = vals[sales_idx] if (sales_idx < len(vals)) else None
                sales = parse_numeric(raw_val)
                if asin not in asin_best or sales > asin_best[asin][0]:
                    asin_best[asin] = (sales, vals)
            sorted_rows = sorted(asin_best.values(), key=lambda x: x[0], reverse=True)
            top = [v[1] for v in sorted_rows[:10]]
            bottom = [v[1] for v in sorted_rows[-10:]] if len(sorted_rows) > 10 else [v[1] for v in sorted_rows]
            total = len(asin_best)
        else:
            def get_sales(item):
                vals = item[1]
                raw_val = vals[sales_idx] if (sales_idx < len(vals)) else None
                return parse_numeric(raw_val)
            sorted_items = sorted(items, key=get_sales, reverse=True)
            top = [x[1] for x in sorted_items[:10]]
            bottom = [x[1] for x in sorted_items[-10:]] if len(sorted_items) > 10 else [x[1] for x in sorted_items]
            total = len(items)
        summary.append((cat, total, len(top), len(bottom)))
        for vals in top:
            best_rows.append((cat, vals))
        for vals in bottom:
            worst_rows.append((cat, vals))
    return best_rows, worst_rows, summary


# ---------------------------------------------------------------------------
# Flask Routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/inspect", methods=["POST"])
def inspect_file():
    """Atomic endpoint: Fast read-only inspection of sheets and column headers (<200ms)."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are supported"}), 400

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name

        try:
            # Load workbook in read_only mode for instant inspection
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            sheet_columns = {}

            for sheet in sheets:
                ws = wb[sheet]
                header_row_idx, raw_headers = detect_header_row(ws)
                headers = build_headers_list(raw_headers)
                sheet_columns[sheet] = {
                    "columns": headers,
                    "header_row": header_row_idx,
                    "keyword_default": auto_pick_keyword_col(headers),
                    "portfolio_default": auto_pick_portfolio_col(headers),
                }
            wb.close()
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

        default_sheet = sheets[0]
        for preferred in ["Exact_Matches_Sorted", "P1_Exact_Candidates", "Sheet1"]:
            if preferred in sheets:
                default_sheet = preferred
                break

        return jsonify({
            "success": True,
            "filename": f.filename,
            "sheets": sheets,
            "default_sheet": default_sheet,
            "sheet_columns": sheet_columns
        })
    except Exception as e:
        import traceback
        return jsonify({"error": f"Cannot read Excel file: {str(e)}", "trace": traceback.format_exc()}), 400


@app.route("/api/process", methods=["POST"])
def process():
    """Atomic endpoint: Processes uploaded file and options in a single HTTP POST request."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded in request."}), 400
    f = request.files["file"]
    sheet_name = request.form.get("sheet")
    keyword_col = request.form.get("keyword_col")
    portfolio_col = request.form.get("portfolio_col")
    analysis_mode = request.form.get("analysis_mode", "Search Term")
    extract_asins_flag = request.form.get("extract_asins", "true").lower() in ("true", "1", "yes")

    if not sheet_name:
        return jsonify({"error": "Please select a target sheet."}), 400
    if not keyword_col:
        return jsonify({"error": "Please select a Keyword column."}), 400

    start_time = time.time()
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name

        try:
            wb_in = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            if sheet_name not in wb_in.sheetnames:
                return jsonify({"error": f"Sheet '{sheet_name}' not found in workbook."}), 400

            ws = wb_in[sheet_name]
            header_row_idx, raw_headers = detect_header_row(ws)
            headers = build_headers_list(raw_headers)
            num_cols = len(headers)
            
            # Reopen the workbook to reset the read_only iterator before the second pass
            wb_in.close()
            wb_in = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb_in[sheet_name]

            try:
                keyword_col_idx = headers.index(keyword_col)
            except ValueError:
                return jsonify({"error": f"Column '{keyword_col}' not found in sheet headers."}), 400

            portfolio_idx = None
            if portfolio_col and portfolio_col != "(None)":
                try:
                    portfolio_idx = headers.index(portfolio_col)
                except ValueError:
                    pass

            data_rows = []
            for row_idx, row_vals in enumerate(
                ws.iter_rows(min_row=header_row_idx + 1, max_col=num_cols, values_only=True),
                start=header_row_idx + 1
            ):
                if not row_vals:
                    continue
                vals = list(row_vals[:num_cols])
                if len(vals) < num_cols:
                    vals.extend([None] * (num_cols - len(vals)))
                
                cell_kw = vals[keyword_col_idx] if keyword_col_idx < len(vals) else None
                cell_first = vals[0] if len(vals) > 0 else None
                if cell_kw and str(cell_kw).strip() != "" and "|" not in str(cell_first or ""):
                    data_rows.append((row_idx, vals))
                    
            wb_in.close()
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

        # ASIN Extraction
        all_asins = set()
        asin_rows_list = []
        if extract_asins_flag:
            for _, vals in data_rows:
                st = str(vals[keyword_col_idx]).strip() if (keyword_col_idx < len(vals) and vals[keyword_col_idx] is not None) else ""
                port = str(vals[portfolio_idx]).strip() if (portfolio_idx is not None and portfolio_idx < len(vals) and vals[portfolio_idx] is not None) else ""
                for asin in extract_asins_from_text(st):
                    if asin not in all_asins:
                        all_asins.add(asin)
                        asin_rows_list.append((st, asin, port))

        # Classify
        classified = []
        for _, vals in data_rows:
            broad_cat = str(vals[0]).strip() if (len(vals) > 0 and vals[0] is not None) else ""
            search_term = str(vals[keyword_col_idx]).strip() if (keyword_col_idx < len(vals) and vals[keyword_col_idx] is not None) else ""
            base_subcat = classify_search_term(broad_cat, search_term)
            color = extract_color(search_term)
            final_subcat = base_subcat
            if color:
                final_subcat += f" - {color}"
            classified.append((final_subcat, search_term, vals))

        # Sales column index detection
        sales_idx = find_sales_column_index(headers)

        # Create output workbook
        wb_out = openpyxl.Workbook()
        if "Sheet" in wb_out.sheetnames:
            del wb_out["Sheet"]
            
        # Recreate the original data sheet to preserve it
        ws_orig = wb_out.create_sheet(sheet_name)
        
        # Write headers
        for c_idx, h in enumerate(headers, 1):
            cell = ws_orig.cell(row=1, column=c_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
        # Write data rows
        for r_idx, (orig_r, vals) in enumerate(data_rows, 2):
            for c_idx, val in enumerate(vals, 1):
                ws_orig.cell(row=r_idx, column=c_idx, value=val)
                
        # Optional: Auto-filter for original data sheet
        if data_rows:
            ws_orig.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(num_cols)}{len(data_rows)+1}"

        sheets_written = [sheet_name]
        categories_found = set()
        do_search = analysis_mode in ("Search Term", "Both")
        do_asin = analysis_mode in ("ASIN", "Both")

        if do_search:
            st_groups = defaultdict(list)
            for subcat, search_term, vals in classified:
                st_groups[subcat].append((search_term, vals))
            best, worst, summary = build_best_worst(st_groups, sales_idx, "Search Term")
            write_analysis_sheet(wb_out, "SearchTerm_Analysis",
                "Search Term Analysis  |  Top 10 Best & Worst per Category",
                "2F5496", headers, "Product Subcategory", best, worst)
            sheets_written.append("SearchTerm_Analysis")
            categories_found.update(st_groups.keys())

        if do_asin:
            asin_groups = defaultdict(list)
            for subcat, search_term, vals in classified:
                asins = extract_asins_from_text(search_term)
                if not asins:
                    continue
                if portfolio_idx is not None and portfolio_idx < len(vals) and vals[portfolio_idx] is not None and str(vals[portfolio_idx]).strip():
                    grp_key = str(vals[portfolio_idx]).strip()
                else:
                    grp_key = subcat
                for asin in asins:
                    asin_groups[grp_key].append((asin, vals))
            cat_label = "Portfolio" if portfolio_idx is not None else "Product Subcategory"
            best, worst, summary = build_best_worst(asin_groups, sales_idx, "ASIN")
            write_analysis_sheet(wb_out, "ASIN_Analysis",
                "ASIN Analysis  |  Top 10 Best & Worst per Category",
                "1B5E20", headers, cat_label, best, worst)
            sheets_written.append("ASIN_Analysis")

        asin_count = 0
        if extract_asins_flag and asin_rows_list:
            write_asin_sheet(wb_out, all_asins, asin_rows_list, portfolio_idx)
            sheets_written.append("Extracted_ASINs")
            asin_count = len(all_asins)

        out_id = str(uuid.uuid4())
        out_path = os.path.join(OUTPUT_DIR, f"{out_id}_output.xlsx")
        wb_out.save(out_path)
        elapsed = time.time() - start_time

        return jsonify({
            "success": True,
            "download_id": out_id,
            "rows_processed": len(data_rows),
            "categories": len(categories_found),
            "sheets_created": sheets_written,
            "asins_extracted": asin_count,
            "elapsed": round(elapsed, 2),
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/download/<download_id>")
def download(download_id):
    out_path = os.path.join(OUTPUT_DIR, f"{download_id}_output.xlsx")
    if not os.path.exists(out_path):
        return jsonify({"error": "Download file expired or not found."}), 404
    return send_file(out_path, as_attachment=True,
                     download_name="SearchTerm_Analysis_Report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
